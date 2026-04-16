"""Excel and PDF exporters for the categorized bill summary."""
from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Iterable, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


@dataclass
class LineResult:
    raw: str
    matched_name: Optional[str]
    department: Optional[str]
    category: str
    amount: Optional[float]
    score: int  # 0-100; 100 = exact


@dataclass
class CategorizedBill:
    bill_name: str
    lines: List[LineResult]
    categories: List[str]

    def raw_totals(self) -> Dict[str, float]:
        """Totals before the even-split of uncategorized charges."""
        t = {c: 0.0 for c in self.categories}
        t.setdefault("Uncategorized", 0.0)
        for ln in self.lines:
            if ln.amount is None:
                continue
            t.setdefault(ln.category, 0.0)
            t[ln.category] += ln.amount
        return t

    def totals(self) -> Dict[str, float]:
        """Final totals: any Uncategorized charges are split evenly across the
        main categories so nothing is left in an "Uncategorized" bucket."""
        raw = self.raw_totals()
        uncat = raw.pop("Uncategorized", 0.0)
        n = len(self.categories)
        if n > 0 and uncat != 0.0:
            per_cat = round(uncat / n, 2)
            # Correct rounding so sum stays exact.
            remainder = round(uncat - per_cat * n, 2)
            for i, cat in enumerate(self.categories):
                raw.setdefault(cat, 0.0)
                raw[cat] += per_cat
                if i == 0:
                    raw[cat] += remainder  # toss leftover penny into first cat
            # Indicate there was a split
            raw["_overhead_split"] = uncat  # stash for display (not a real category)
        return raw

    def grand_total(self) -> float:
        return sum(ln.amount for ln in self.lines if ln.amount is not None)

    def unmatched(self) -> List[LineResult]:
        return [ln for ln in self.lines if not ln.matched_name]


# ---- Excel --------------------------------------------------------------

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def export_excel(bill: CategorizedBill, out_path: str) -> str:
    wb = openpyxl.Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Bill Categorization Summary"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Source: {bill.bill_name}"
    ws["A3"] = f"Generated: {datetime.now():%Y-%m-%d %H:%M:%S}"

    ws["A5"] = "Category"
    ws["B5"] = "Total"
    for c in ("A5", "B5"):
        ws[c].fill = _HEADER_FILL
        ws[c].font = _HEADER_FONT
        ws[c].border = _BORDER

    row = 6
    totals = bill.totals()
    overhead = totals.pop("_overhead_split", 0.0)
    for cat in bill.categories:
        ws.cell(row=row, column=1, value=cat).border = _BORDER
        c = ws.cell(row=row, column=2, value=round(totals.get(cat, 0.0), 2))
        c.number_format = '"$"#,##0.00'
        c.border = _BORDER
        row += 1
    ws.cell(row=row, column=1, value="Grand Total").font = Font(bold=True)
    gt = ws.cell(row=row, column=2, value=round(bill.grand_total(), 2))
    gt.font = Font(bold=True)
    gt.number_format = '"$"#,##0.00'
    if overhead:
        row += 1
        ws.cell(row=row, column=1, value=f"  (Incl. ${overhead:,.2f} overhead split ÷ {len(bill.categories)})").font = Font(italic=True, color="808080")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # Detail sheet
    ds = wb.create_sheet("Line Items")
    headers = ["Row", "Matched Employee", "Department", "Category", "Amount", "Match Score", "Raw Line"]
    for i, h in enumerate(headers, start=1):
        c = ds.cell(row=1, column=i, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.border = _BORDER
    for i, ln in enumerate(bill.lines, start=2):
        ds.cell(row=i, column=1, value=i - 1)
        ds.cell(row=i, column=2, value=ln.matched_name or "")
        ds.cell(row=i, column=3, value=ln.department or "")
        ds.cell(row=i, column=4, value=ln.category)
        ac = ds.cell(row=i, column=5, value=round(ln.amount, 2) if ln.amount is not None else "")
        ac.number_format = '"$"#,##0.00'
        ds.cell(row=i, column=6, value=ln.score)
        ds.cell(row=i, column=7, value=ln.raw[:500])
    widths = [6, 28, 28, 16, 14, 12, 80]
    for i, w in enumerate(widths, start=1):
        ds.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ds.freeze_panes = "A2"

    # Unmatched sheet
    unmatched = bill.unmatched()
    if unmatched:
        us = wb.create_sheet("Unmatched")
        for i, h in enumerate(["Row", "Amount", "Raw Line"], start=1):
            c = us.cell(row=1, column=i, value=h)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
        for i, ln in enumerate(unmatched, start=2):
            us.cell(row=i, column=1, value=i - 1)
            ac = us.cell(row=i, column=2, value=round(ln.amount, 2) if ln.amount is not None else "")
            ac.number_format = '"$"#,##0.00'
            us.cell(row=i, column=3, value=ln.raw[:500])
        us.column_dimensions["A"].width = 6
        us.column_dimensions["B"].width = 14
        us.column_dimensions["C"].width = 100

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


# ---- AP Phone Line Analysis (auto-generated) -----------------------------

def export_ap_analysis(bill: CategorizedBill, out_path: str) -> str:
    """Simplified phone-line breakdown for Accounts Payable.

    Columns: Phone | Employee | EE Dept | Category | Amount
    Sorted by category then amount descending.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phone Line Analysis"

    headers = ["Phone", "Employee", "EE Dept", "Category", "Amount"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center")

    # Build rows — only lines that have a phone number are relevant here.
    rows_data = []
    for ln in bill.lines:
        # Derive phone from the raw text if it was a phone-based line.
        phone_str = ""
        if hasattr(ln, "raw") and ln.raw:
            import re
            m = re.search(r"\(\d{3}\)\s*\d{3}[\-\s.]\d{4}", ln.raw)
            if m:
                phone_str = m.group(0)
        rows_data.append((
            phone_str,
            ln.matched_name or "",
            ln.department or "",
            ln.category,
            ln.amount,
        ))

    rows_data.sort(key=lambda r: (r[3], -(r[4] or 0)))

    for i, row in enumerate(rows_data, 2):
        ws.cell(row=i, column=1, value=row[0]).border = _BORDER
        ws.cell(row=i, column=2, value=row[1]).border = _BORDER
        ws.cell(row=i, column=3, value=row[2]).border = _BORDER
        ws.cell(row=i, column=4, value=row[3]).border = _BORDER
        ac = ws.cell(row=i, column=5, value=round(row[4], 2) if row[4] is not None else "")
        ac.number_format = '"$"#,##0.00'
        ac.border = _BORDER

    # Summary at the bottom
    last = len(rows_data) + 2
    ws.cell(row=last, column=3, value="").border = _BORDER
    ws.cell(row=last, column=4, value="Grand Total").font = Font(bold=True)
    ws.cell(row=last, column=4).border = _BORDER
    gt = ws.cell(row=last, column=5, value=round(bill.grand_total(), 2))
    gt.font = Font(bold=True)
    gt.number_format = '"$"#,##0.00'
    gt.border = _BORDER

    # Category subtotals
    last += 1
    ws.cell(row=last, column=1, value="").border = _BORDER
    last += 1
    ws.cell(row=last, column=1, value="Category Totals").font = Font(bold=True, size=11)
    totals = bill.totals()
    overhead = totals.pop("_overhead_split", 0.0)
    last += 1
    for cat in bill.categories:
        ws.cell(row=last, column=4, value=cat).border = _BORDER
        c = ws.cell(row=last, column=5, value=round(totals.get(cat, 0.0), 2))
        c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True)
        c.border = _BORDER
        last += 1
    if overhead:
        ws.cell(row=last, column=4, value=f"(Incl. ${overhead:,.2f} overhead ÷ {len(bill.categories)})").font = Font(italic=True, color="808080")

    widths = [16, 28, 26, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(rows_data) + 1}"

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


# ---- PDF ----------------------------------------------------------------

def export_pdf(bill: CategorizedBill, out_path: str) -> str:
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    doc = SimpleDocTemplate(
        out_path,
        pagesize=LETTER,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()
    elements = []

    # Optional logo header
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "logo.png")
    if os.path.exists(logo_path):
        try:
            elements.append(RLImage(logo_path, width=2.4 * inch, height=0.7 * inch))
            elements.append(Spacer(1, 0.1 * inch))
        except Exception:
            pass

    elements.append(Paragraph("Bill Categorization Summary", styles["Title"]))
    elements.append(Paragraph(f"Source: {bill.bill_name}", styles["Normal"]))
    elements.append(Paragraph(f"Generated: {datetime.now():%Y-%m-%d %H:%M:%S}", styles["Normal"]))
    elements.append(Spacer(1, 0.25 * inch))

    totals = bill.totals()
    overhead = totals.pop("_overhead_split", 0.0)
    data = [["Category", "Total"]]
    for cat in bill.categories:
        data.append([cat, f"${totals.get(cat, 0.0):,.2f}"])
    data.append(["Grand Total", f"${bill.grand_total():,.2f}"])
    if overhead:
        data.append([f"(Incl. ${overhead:,.2f} overhead split ÷ {len(bill.categories)})", ""])

    t = Table(data, colWidths=[3.5 * inch, 2 * inch])
    t.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9E1F2")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
        ])
    )
    elements.append(t)
    elements.append(Spacer(1, 0.3 * inch))

    # Line items table
    elements.append(Paragraph("Line Items", styles["Heading2"]))
    line_rows = [["Employee", "Dept", "Category", "Amount"]]
    for ln in bill.lines:
        line_rows.append([
            (ln.matched_name or "—")[:30],
            (ln.department or "—")[:25],
            ln.category,
            f"${ln.amount:,.2f}" if ln.amount is not None else "—",
        ])
    lt = Table(line_rows, colWidths=[2.2 * inch, 1.8 * inch, 1.3 * inch, 1.0 * inch], repeatRows=1)
    lt.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ])
    )
    elements.append(lt)

    unmatched = bill.unmatched()
    if unmatched:
        elements.append(Spacer(1, 0.25 * inch))
        elements.append(Paragraph(f"Unmatched ({len(unmatched)})", styles["Heading2"]))
        elements.append(Paragraph(
            "These rows could not be tied to an employee. Review and reassign manually.",
            styles["Italic"],
        ))
        u_rows = [["Amount", "Raw Line"]]
        for ln in unmatched:
            u_rows.append([
                f"${ln.amount:,.2f}" if ln.amount is not None else "—",
                ln.raw[:80],
            ])
        ut = Table(u_rows, colWidths=[1.0 * inch, 5.3 * inch], repeatRows=1)
        ut.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C00000")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 1), (0, -1), "RIGHT"),
            ])
        )
        elements.append(ut)

    doc.build(elements)
    return out_path
